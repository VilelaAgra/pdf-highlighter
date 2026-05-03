import fitz
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def extrair_valor(texto: str) -> dict:
    # Captura todos os padrões "R$ valor" da linha (com ou sem sinal)
    # O primeiro é sempre a transação, o segundo é o saldo — ignoramos o saldo
    ocorrencias = re.findall(r'(-\s*)?R\$\s*([\d.,]+)', texto)

    if ocorrencias:
        sinal_str, num_str = ocorrencias[0]
        try:
            numero = float(num_str.replace('.', '').replace(',', '.'))
            is_negativo = bool(sinal_str.strip())
            return {
                "valor": -numero if is_negativo else numero,
                "sinal_inferido": False
            }
        except:
            pass

    return {"valor": None, "sinal_inferido": True}

def hex_para_rgb(hex_color: str) -> tuple[float, float, float]:
    hex_color = hex_color.lstrip("#")
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    return (r / 255, g / 255, b / 255)

def processar_pdf(pdf_bytes: bytes, filtros: list[dict]) -> tuple[bytes, dict]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    relatorio = {
        "total_linhas_marcadas": 0,
        "paginas_afetadas": [],
        "soma_valores": 0.0,
        "grupos": {
            f["label"]: {
                "total": 0,
                "soma": 0.0,
                "total_creditos": 0.0,
                "total_debitos": 0.0,
                "cor": f["cor"],
                "linhas": []
            } for f in filtros
        }
    }

    for num_pagina in range(len(doc)):
        pagina = doc[num_pagina]
        blocos = pagina.get_text("dict")["blocks"]
        largura_pagina = pagina.rect.width
        pagina_teve_match = False

        for bloco in blocos:
            if bloco.get("type") != 0:
                continue

            texto_bloco = " ".join(
                span["text"]
                for linha in bloco.get("lines", [])
                for span in linha.get("spans", [])
            ).strip()

            if not texto_bloco:
                continue

            for filtro in filtros:
                keywords = filtro["keywords"]
                cor_rgb = hex_para_rgb(filtro["cor"])

                if any(kw.lower() in texto_bloco.lower() for kw in keywords):
                    bbox = fitz.Rect(bloco["bbox"])
                    bbox_completo = fitz.Rect(0, bbox.y0 - 2, largura_pagina, bbox.y1 + 2)
                    pagina.draw_rect(bbox_completo, color=cor_rgb, fill=cor_rgb, fill_opacity=0.4)

                    resultado_valor = extrair_valor(texto_bloco)
                    valor = resultado_valor["valor"]
                    sinal_inferido = resultado_valor["sinal_inferido"]
                    label = filtro["label"]

                    relatorio["grupos"][label]["linhas"].append({
                        "pagina": num_pagina + 1,
                        "texto": texto_bloco,
                        "valor_detectado": valor,
                        "sinal_inferido": sinal_inferido
                    })
                    relatorio["grupos"][label]["total"] += 1

                    if valor is not None:
                        relatorio["grupos"][label]["soma"] += valor
                        relatorio["soma_valores"] += valor
                        if valor >= 0:
                            relatorio["grupos"][label]["total_creditos"] += valor
                        else:
                            relatorio["grupos"][label]["total_debitos"] += valor

                    pagina_teve_match = True
                    break

        if pagina_teve_match and (num_pagina + 1) not in relatorio["paginas_afetadas"]:
            relatorio["paginas_afetadas"].append(num_pagina + 1)

    pdf_marcado = doc.tobytes()
    doc.close()
    return pdf_marcado, relatorio

def gerar_excel(relatorio: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.column_dimensions["A"].width = 20
    ws_resumo.column_dimensions["B"].width = 15
    ws_resumo.column_dimensions["C"].width = 20
    ws_resumo.column_dimensions["D"].width = 20
    ws_resumo.column_dimensions["E"].width = 20

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    cabecalhos = ["Grupo", "Linhas", "Créditos", "Débitos", "Saldo líquido"]
    for col, titulo in enumerate(cabecalhos, start=1):
        cell = ws_resumo.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    def fmt(v: float) -> str:
        return f'R$ {v:_.2f}'.replace(".", ",").replace("_", ".")

    for row, (label, grupo) in enumerate(relatorio["grupos"].items(), start=2):
        cor_hex = grupo["cor"].lstrip("#")
        grupo_fill = PatternFill("solid", fgColor=cor_hex)

        ws_resumo.cell(row=row, column=1, value=label).fill = grupo_fill
        ws_resumo.cell(row=row, column=2, value=grupo["total"]).fill = grupo_fill
        ws_resumo.cell(row=row, column=3, value=fmt(grupo["total_creditos"])).fill = grupo_fill
        ws_resumo.cell(row=row, column=4, value=fmt(grupo["total_debitos"])).fill = grupo_fill
        ws_resumo.cell(row=row, column=5, value=fmt(grupo["soma"])).fill = grupo_fill

    for label, grupo in relatorio["grupos"].items():
        ws = wb.create_sheet(label[:31])
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 18

        cabecalhos_detalhe = ["Página", "Texto", "Valor detectado", "Sinal inferido?"]
        for col, titulo in enumerate(cabecalhos_detalhe, start=1):
            cell = ws.cell(row=1, column=col, value=titulo)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        cor_hex = grupo["cor"].lstrip("#")
        linha_fill = PatternFill("solid", fgColor=cor_hex)
        alerta_fill = PatternFill("solid", fgColor="FEF08A")  # amarelo alerta

        for row, linha in enumerate(grupo["linhas"], start=2):
            fill = alerta_fill if linha["sinal_inferido"] else linha_fill
            ws.cell(row=row, column=1, value=linha["pagina"]).fill = fill
            ws.cell(row=row, column=2, value=linha["texto"]).fill = fill
            valor_fmt = fmt(linha["valor_detectado"]) if linha["valor_detectado"] is not None else "—"
            ws.cell(row=row, column=3, value=valor_fmt).fill = fill
            ws.cell(row=row, column=4, value="⚠️ Conferir" if linha["sinal_inferido"] else "✓").fill = fill

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()